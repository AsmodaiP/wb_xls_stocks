from asyncio import FIRST_EXCEPTION
from datetime import datetime
from itertools import count
from time import time
import openpyxl
from collections import Counter
from openpyxl.styles import PatternFill

FIRST_INDEX = 5


def get_counts_from_table(name='db.xlsx'):
    wb = openpyxl.load_workbook(name)
    if name != 'поставка.xlsx': 
        wb.active = 2
    ws = wb.active
    values = []
    for row in ws.rows:

        try:
            value = str(row[0].value)
            if value.isdigit():
                values.append(value)
        except AttributeError:
            pass
    cnt = Counter(values)
    return cnt


def barcode_in_data(barcode, data):
    if barcode is None or barcode == 'Бакрод':
        return False
    if str(barcode) not in data:
        return False
    return True


def get_mirrors(wb):

    wb.active = len(wb.sheetnames) - 1
    ws = wb.active
    mirrors = []
    for row in ws.rows:
        if row[0].value is None:
            continue
        val = row[0].value.replace(' ', '')
        mirrors.append(val.split(','))
    return mirrors


def get_all_mirrors(barcode, mirrors):
    for mirors_barcodes in mirrors:
        if str(barcode) in mirors_barcodes:
            return set(mirors_barcodes)
    return set([str(barcode)])


def insert_data_in_table(data, first_index,fill, target_table='stocks.xlsx'):
    wb = openpyxl.load_workbook(target_table)
    mirrors = get_mirrors(wb)
    for list_number in range(len(wb.sheetnames) - 1):
        wb.active = list_number
        ws = wb.active
        for row in ws.rows:
            barcode = row[0].value
            mirrors_barcodes = get_all_mirrors(barcode, mirrors)
            new_value = get_sum_by_barcode(mirrors_barcodes, data, row)
            if new_value != 0:
                row[first_index + datetime.now().day].value = new_value
                row[first_index + datetime.now().day].fill = fill
        wb.save(target_table)
    return {'erorrs': data}


def get_sum_by_barcode(mirrors_barcodes, data, row):
    sum_of_mirrors = 0
    for bar in mirrors_barcodes:
        if str(bar) in data:
            sum_of_mirrors += data[str(bar)]
            del data[str(bar)]

    today_sells = row[FIRST_INDEX + datetime.now().day].value
    if today_sells is not None:
        return sum([int(today_sells), sum_of_mirrors])
    return sum_of_mirrors


def insert_sales(filename):
    fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    return insert_data_in_table(
        get_counts_from_table(filename), first_index=FIRST_INDEX, fill=fill)

def insert_supplie(filename):
    fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    return insert_data_in_table(
        get_counts_from_table(filename), first_index=FIRST_INDEX+33, fill=fill) 

# 2000393474004
if __name__ == '__main__':
    print(get_counts_from_table())